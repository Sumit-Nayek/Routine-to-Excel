
# import os
# import json
# import re
# import base64
# from datetime import datetime, date
# from io import BytesIO
# import streamlit as st
# import pandas as pd
# import fitz  # PyMuPDF
# from huggingface_hub import InferenceClient
# import openpyxl
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# st.set_page_config(page_title="Attendance Tracker Generator", page_icon="📝", layout="wide")
# st.title("📝 Attendance & Holiday Tracker Generator")
# st.write("Upload your Class Routine and Academic Holiday Notice (Scanned or Digital PDF).")

# # 1. API Token Setup
# HF_TOKEN = os.getenv("HF_TOKEN")
# if not HF_TOKEN:
#     st.error("HF_TOKEN not detected in environment. Run: export HF_TOKEN='your_token'")
#     st.stop()

# client = InferenceClient(api_key=HF_TOKEN)

# # Month Name to Number Mapper
# MONTH_MAP = {
#     "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
#     "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6,
#     "july": 7, "jul": 7, "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
#     "october": 10, "oct": 10, "november": 11, "nov": 11, "december": 12, "dec": 12
# }

# # Helper: Convert PDF first page to Base64 Image
# def pdf_to_base64_image(uploaded_file):
#     uploaded_file.seek(0)
#     doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
#     page = doc[0]  # First page containing the table
#     pix = page.get_pixmap(dpi=200)  # Render at high resolution for OCR
#     img_bytes = pix.tobytes("png")
#     return base64.b64encode(img_bytes).decode("utf-8")

# # Helper: Extract text from digital PDF (fallback)
# def extract_text_from_pdf(uploaded_file):
#     uploaded_file.seek(0)
#     doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
#     text = ""
#     for page in doc:
#         text += page.get_text() + "\n"
#     return text

# # Helper: Clean subject names
# def clean_subject_name(raw_name: str) -> str:
#     cleaned = re.sub(r'[\(\[\{].*?[\)\]\}]', '', raw_name)
#     cleaned = re.sub(r'^[A-Z]{2,5}[-\s]?\d{2,4}\s*[:\-–]?\s*', '', cleaned, flags=re.IGNORECASE)
#     cleaned = re.sub(r'\s*[:\-–]?\s*[A-Z]{2,5}[-\s]?\d{2,4}$', '', cleaned, flags=re.IGNORECASE)
#     return cleaned.strip(" -:–\t\n")

# # Extract Subjects from Routine
# def extract_routine_subjects(uploaded_file):
#     # Try text extraction first
#     routine_text = extract_text_from_pdf(uploaded_file)
    
#     if len(routine_text.strip()) > 50:
#         # Use fast text model if selectable text exists
#         system_prompt = (
#             "You are an academic routine parser. Extract ONLY the unique, human-readable subject names.\n"
#             "STRICT RULES:\n"
#             "1. Extract full descriptive subject names (e.g., 'Operating Systems', 'Calculus').\n"
#             "2. DO NOT return subject codes (e.g., 'CS301', 'PCC-CS501').\n"
#             "3. Ignore labels like 'Lunch', 'Break', 'Library', 'Mentor Mentee'.\n"
#             "Return JSON ONLY: {\"subjects\": [\"Subject 1\", \"Subject 2\"]}"
#         )
#         response = client.chat.completions.create(
#             model="Qwen/Qwen2.5-Coder-32B-Instruct",
#             messages=[
#                 {"role": "system", "content": system_prompt},
#                 {"role": "user", "content": f"Extract subjects:\n\n{routine_text}"}
#             ],
#             max_tokens=2048,
#             temperature=0.0
#         )
#         raw = response.choices[0].message.content.strip()
#     else:
#         # Fallback to Vision Model if routine is also a scanned image
#         base64_img = pdf_to_base64_image(uploaded_file)
#         response = client.chat.completions.create(
#             model="Qwen/Qwen2.5-VL-72B-Instruct",
#             messages=[
#                 {
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": "Extract all course/subject titles from this schedule image. Exclude codes like CS101. Output JSON: {\"subjects\": [\"Name 1\", \"Name 2\"]}"},
#                         {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_img}"}}
#                     ]
#                 }
#             ],
#             max_tokens=2048,
#             temperature=0.0
#         )
#         raw = response.choices[0].message.content.strip()

#     if "```" in raw:
#         raw = raw.split("```")[1].replace("json", "").strip()

#     parsed = json.loads(raw)
#     cleaned = [clean_subject_name(s) for s in parsed.get("subjects", []) if clean_subject_name(s)]
#     return sorted(list(set(cleaned)))

# # Extract Holidays from Scanned Holiday PDF using Vision LLM
# def extract_holidays_from_pdf_image(uploaded_file, reference_year=2026):
#     base64_img = pdf_to_base64_image(uploaded_file)
    
#     system_prompt = (
#         f"You are an expert document OCR parser. Read the holiday table in this image carefully. "
#         f"The reference year is {reference_year}.\n\n"
#         "CRITICAL INSTRUCTION:\n"
#         "1. Extract ONLY the items listed under the main 'HOLIDAYS:' section table[cite: 1].\n"
#         "2. STRICTLY IGNORE and EXCLUDE the entire 'RESTRICTED HOLIDAYS:' section table[cite: 1]. Do not include any restricted holidays.\n\n"
#         "For each row in the main 'HOLIDAYS' table[cite: 1], extract the 'occasion' (e.g., 'Republic Day') and convert the date (e.g., '26th January') into strict 'YYYY-MM-DD' format (e.g., '2026-01-26')[cite: 1].\n\n"
#         "Return JSON ONLY with this exact structure:\n"
#         "{\n"
#         '  "holidays": [\n'
#         '    {"occasion": "Republic Day", "date": "2026-01-26"},\n'
#         '    {"occasion": "Good Friday", "date": "2026-04-03"}\n'
#         "  ]\n"
#         "}\n"
#         "Do not output any markdown formatting (no ```json). Output raw JSON only."
#     )

#     response = client.chat.completions.create(
#         model="Qwen/Qwen2.5-VL-72B-Instruct",
#         messages=[
#             {
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": system_prompt},
#                     {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_img}"}}
#                 ]
#             }
#         ],
#         max_tokens=3500,
#         temperature=0.0
#     )

#     raw = response.choices[0].message.content.strip()
#     if "```" in raw:
#         raw = raw.split("```")[1].replace("json", "").strip()

#     parsed = json.loads(raw)
    
#     holiday_dict = {}
#     for h in parsed.get("holidays", []):
#         occasion = h.get("occasion", "Holiday").strip()
#         d_str = h.get("date", "").strip()
        
#         try:
#             dt = datetime.strptime(d_str, "%Y-%m-%d").date()
#             holiday_dict[dt.strftime("%Y-%m-%d")] = occasion
#         except ValueError:
#             m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)', d_str)
#             if m and m.group(2).lower() in MONTH_MAP:
#                 dt = datetime(reference_year, MONTH_MAP[m.group(2).lower()], int(m.group(1))).date()
#                 holiday_dict[dt.strftime("%Y-%m-%d")] = occasion

#     return holiday_dict

# # Build Formatted Excel Sheet
# def generate_excel_sheet(subjects, holidays, start_date, end_date):
#     date_range = pd.date_range(start=start_date, end=end_date, freq="B")  # Mon-Fri
    
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Attendance_Tracker"

#     # Styling definitions
#     header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
#     header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
#     holiday_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
#     holiday_font = Font(name="Calibri", size=10, bold=True, color="C00000")
    
#     thin_border = Border(
#         left=Side(style='thin', color='D9D9D9'),
#         right=Side(style='thin', color='D9D9D9'),
#         top=Side(style='thin', color='D9D9D9'),
#         bottom=Side(style='thin', color='D9D9D9')
#     )

#     headers = ["Date", "Day", "Remarks / Holiday"] + subjects
#     ws.append(headers)

#     for col_idx in range(1, len(headers) + 1):
#         cell = ws.cell(row=1, column=col_idx)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     for row_idx, dt in enumerate(date_range, start=2):
#         date_str = dt.strftime("%Y-%m-%d")
#         day_str = dt.strftime("%A")
#         holiday_name = holidays.get(date_str, "")

#         row_data = [date_str, day_str, holiday_name] + [""] * len(subjects)
#         ws.append(row_data)

#         is_holiday = bool(holiday_name)
#         for col_idx in range(1, len(headers) + 1):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             if is_holiday:
#                 cell.fill = holiday_fill
#                 if col_idx == 3:
#                     cell.font = holiday_font
#                     cell.alignment = Alignment(horizontal="left", vertical="center")

#     # Set column widths
#     ws.column_dimensions['A'].width = 14
#     ws.column_dimensions['B'].width = 14
#     ws.column_dimensions['C'].width = 34
#     for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(4, len(headers) + 1)]:
#         ws.column_dimensions[col_letter].width = 24

#     ws.row_dimensions[1].height = 28
#     output = BytesIO()
#     wb.save(output)
#     return output.getvalue()

# # --- Streamlit UI ---
# col_u1, col_u2 = st.columns(2)
# with col_u1:
#     routine_file = st.file_uploader("1. Upload Class Routine (PDF)", type=["pdf"])
# with col_u2:
#     holiday_file = st.file_uploader("2. Upload Holiday Calendar (PDF)", type=["pdf"])

# col_d1, col_d2 = st.columns(2)
# with col_d1:
#     start_input = st.date_input("Semester Start Date", value=date(2026, 1, 1))
# with col_d2:
#     end_input = st.date_input("Semester End Date", value=date(2026, 12, 31))

# if routine_file and st.button("🚀 Generate Attendance Tracker"):
#     if start_input > end_input:
#         st.error("Start Date cannot be after End Date.")
#         st.stop()

#     with st.spinner("Rendering PDF pages and running visual extraction..."):
#         try:
#             # 1. Parse Routine Subjects
#             subjects = extract_routine_subjects(routine_file)
#             if not subjects:
#                 st.error("Could not extract subjects from the routine.")
#                 st.stop()

#             # 2. Parse Scanned Holiday Notice
#             holidays = {}
#             if holiday_file:
#                 holidays = extract_holidays_from_pdf_image(holiday_file, reference_year=start_input.year)
                
#                 with st.expander(f"📅 View Detected Holidays ({len(holidays)} found)", expanded=True):
#                     holiday_df = pd.DataFrame(
#                         [{"Date": d, "Holiday Name": occ} for d, occ in sorted(holidays.items())]
#                     )
#                     st.dataframe(holiday_df, use_container_width=True)

#             # 3. Generate Excel
#             excel_data = generate_excel_sheet(subjects, holidays, start_input, end_input)

#             st.success("Attendance Tracker created successfully!")
#             st.download_button(
#                 label="📥 Download Excel Tracker",
#                 data=excel_data,
#                 file_name=f"attendance_tracker_{start_input.year}.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )

#         except Exception as e:
#             st.error(f"Error during processing: {e}")
import os
import json
import re
from datetime import datetime, date
from io import BytesIO
import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
from google import genai
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="Attendance, Holiday & Exam Tracker", page_icon="📝", layout="wide")
st.title("📝 Student Attendance, Holiday & Exam Tracker Generator")
st.write("Upload your Class Routine, Holiday Calendar, and Exam Schedule (PDFs) to create an automated attendance sheet.")

# 1. Initialize GenAI Client
GEMINI_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_KEY:
    st.error("GEMINI_API_KEY not found in environment. Run: export GEMINI_API_KEY='your_key'")
    st.stop()

client = genai.Client(api_key=GEMINI_KEY)

# Model fallback pool to prevent 503 traffic spike errors
FALLBACK_MODELS = [
    "gemini-3.7-flash",
    "gemini-3.5-flash",
    "gemini-flash-latest",
    "gemini-flash-lite-latest",
    "gemini-pro-latest"
]

def generate_with_fallback(contents):
    last_error = None
    for model_name in FALLBACK_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=contents
            )
            return response.text.strip()
        except Exception as e:
            last_error = e
            continue
    raise last_error

MONTH_MAP = {
    "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
    "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6,
    "july": 7, "jul": 7, "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10, "november": 11, "nov": 11, "december": 12, "dec": 12
}

# Helper: Convert PDF page to PIL Image
def pdf_to_pil_image(uploaded_file, page_idx=0):
    uploaded_file.seek(0)
    doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
    page = doc[page_idx]
    pix = page.get_pixmap(dpi=200)
    return Image.open(BytesIO(pix.tobytes("png")))

# Helper: Clean subject names
def clean_subject_name(raw_name: str) -> str:
    cleaned = re.sub(r'[\(\[\{].*?[\)\]\}]', '', raw_name)
    cleaned = re.sub(r'^[A-Z]{2,5}[-\s]?\d{2,4}\s*[:\-–]?\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s*[:\-–]?\s*[A-Z]{2,5}[-\s]?\d{2,4}$', '', cleaned, flags=re.IGNORECASE)
    return cleaned.strip(" -:–\t\n")

# 1. Extract Subjects from Class Routine
def extract_routine_subjects(uploaded_file):
    img = pdf_to_pil_image(uploaded_file)
    prompt = (
        "Extract all unique, human-readable course/subject titles from this timetable image.\n"
        "STRICT RULES:\n"
        "1. Extract ONLY human-readable subject names (e.g. 'Operating Systems', 'Calculus').\n"
        "2. EXCLUDE subject codes (like CS101, PCC-CS501).\n"
        "3. EXCLUDE 'Lunch', 'Break', 'Library', 'Mentor Mentee'.\n"
        "Return ONLY a JSON object: {\"subjects\": [\"Subject 1\", \"Subject 2\"]}"
    )
    raw = generate_with_fallback([img, prompt])
    if "```" in raw:
        raw = raw.split("```")[1].replace("json", "").strip()

    parsed = json.loads(raw)
    cleaned = [clean_subject_name(s) for s in parsed.get("subjects", []) if clean_subject_name(s)]
    return sorted(list(set(cleaned)))

# 2. Extract ONLY Main Holidays
def extract_holidays_from_pdf(uploaded_file, reference_year=2026):
    img = pdf_to_pil_image(uploaded_file)
    prompt = (
        f"You are an expert document OCR parser. Read the holiday notice image carefully. Reference year: {reference_year}.\n"
        "RULES:\n"
        "1. Extract ONLY entries from the main 'HOLIDAYS:' section table.\n"
        "2. STRICTLY EXCLUDE any rows from the 'RESTRICTED HOLIDAYS:' table.\n"
        "3. Convert every extracted date to 'YYYY-MM-DD' format.\n\n"
        "Return ONLY a JSON object: {\"holidays\": [{\"occasion\": \"Republic Day\", \"date\": \"2026-01-26\"}]}"
    )
    raw = generate_with_fallback([img, prompt])
    if "```" in raw:
        raw = raw.split("```")[1].replace("json", "").strip()

    parsed = json.loads(raw)
    holiday_dict = {}
    for h in parsed.get("holidays", []):
        occasion = h.get("occasion", "Holiday").strip()
        d_str = h.get("date", "").strip()
        try:
            dt = datetime.strptime(d_str, "%Y-%m-%d").date()
            holiday_dict[dt.strftime("%Y-%m-%d")] = occasion
        except ValueError:
            m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)', d_str)
            if m and m.group(2).lower() in MONTH_MAP:
                dt = datetime(reference_year, MONTH_MAP[m.group(2).lower()], int(m.group(1))).date()
                holiday_dict[dt.strftime("%Y-%m-%d")] = occasion
    return holiday_dict

# 3. Extract Exam Dates and Papers from Exam Schedule
def extract_exams_from_pdf(uploaded_file, reference_year=2026):
    img = pdf_to_pil_image(uploaded_file)
    prompt = (
        f"You are an exam schedule parser. Read the exam timetable image carefully. Reference year: {reference_year}.\n"
        "RULES:\n"
        "1. Extract all exam dates and corresponding subject/paper names.\n"
        "2. Convert every date to 'YYYY-MM-DD' format.\n"
        "3. Format each event as 'Exam: <Subject Name>'.\n\n"
        "Return ONLY a JSON object: {\"exams\": [{\"subject\": \"Operating Systems\", \"date\": \"2026-05-12\"}]}"
    )
    raw = generate_with_fallback([img, prompt])
    if "```" in raw:
        raw = raw.split("```")[1].replace("json", "").strip()

    parsed = json.loads(raw)
    exam_dict = {}
    for e in parsed.get("exams", []):
        subj = clean_subject_name(e.get("subject", "Mid-term/Final Exam"))
        d_str = e.get("date", "").strip()
        try:
            dt = datetime.strptime(d_str, "%Y-%m-%d").date()
            exam_dict[dt.strftime("%Y-%m-%d")] = f"Exam: {subj}"
        except ValueError:
            m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)', d_str)
            if m and m.group(2).lower() in MONTH_MAP:
                dt = datetime(reference_year, MONTH_MAP[m.group(2).lower()], int(m.group(1))).date()
                exam_dict[dt.strftime("%Y-%m-%d")] = f"Exam: {subj}"
    return exam_dict

# 4. Excel Builder with Custom Color Coding
def generate_excel_sheet(subjects, holidays, exams, start_date, end_date):
    date_range = pd.date_range(start=start_date, end=end_date, freq="B")  # Mon-Fri
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance_Tracker"

    # Color Fills & Fonts
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Navy Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    holiday_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Soft Coral/Peach
    holiday_font = Font(name="Calibri", size=10, bold=True, color="C00000")

    exam_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft Gold/Amber
    exam_font = Font(name="Calibri", size=10, bold=True, color="B25900")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = ["Date", "Day", "Remarks / Events"] + subjects
    ws.append(headers)

    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Populate Data Rows
    for row_idx, dt in enumerate(date_range, start=2):
        date_str = dt.strftime("%Y-%m-%d")
        day_str = dt.strftime("%A")

        is_exam = date_str in exams
        is_holiday = date_str in holidays

        # Priority: Exam > Holiday
        if is_exam:
            remark = exams[date_str]
        elif is_holiday:
            remark = holidays[date_str]
        else:
            remark = ""

        row_data = [date_str, day_str, remark] + [""] * len(subjects)
        ws.append(row_data)

        # Style Cells
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if is_exam:
                cell.fill = exam_fill
                if col_idx == 3:
                    cell.font = exam_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            elif is_holiday:
                cell.fill = holiday_fill
                if col_idx == 3:
                    cell.font = holiday_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column Widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 36
    for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(4, len(headers) + 1)]:
        ws.column_dimensions[col_letter].width = 24

    ws.row_dimensions[1].height = 28
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# --- Streamlit UI Layout ---
col_u1, col_u2, col_u3 = st.columns(3)
with col_u1:
    routine_file = st.file_uploader("1. Class Routine (PDF)", type=["pdf"])
with col_u2:
    holiday_file = st.file_uploader("2. Holiday Notice (PDF, Optional)", type=["pdf"])
with col_u3:
    exam_file = st.file_uploader("3. Exam Schedule (PDF, Optional)", type=["pdf"])

col_d1, col_d2 = st.columns(2)
with col_d1:
    start_input = st.date_input("Semester Start Date", value=date(2026, 1, 1))
with col_d2:
    end_input = st.date_input("Semester End Date", value=date(2026, 12, 31))

if routine_file and st.button("🚀 Generate Attendance Tracker"):
    if start_input > end_input:
        st.error("Start Date cannot be after End Date.")
        st.stop()

    with st.spinner("Analyzing schedule, holidays, and exam dates..."):
        try:
            # 1. Subjects
            subjects = extract_routine_subjects(routine_file)
            if not subjects:
                st.error("Could not extract subjects from the routine.")
                st.stop()

            # 2. Holidays
            holidays = {}
            if holiday_file:
                holidays = extract_holidays_from_pdf(holiday_file, reference_year=start_input.year)
                with st.expander(f"📅 View Detected Holidays ({len(holidays)} found)"):
                    st.dataframe(pd.DataFrame([{"Date": d, "Holiday": h} for d, h in sorted(holidays.items())]), use_container_width=True)

            # 3. Exams
            exams = {}
            if exam_file:
                exams = extract_exams_from_pdf(exam_file, reference_year=start_input.year)
                with st.expander(f"📝 View Detected Exams ({len(exams)} found)"):
                    st.dataframe(pd.DataFrame([{"Date": d, "Exam": e} for d, e in sorted(exams.items())]), use_container_width=True)

            # 4. Generate Sheet
            excel_data = generate_excel_sheet(subjects, holidays, exams, start_input, end_input)

            st.success("Attendance Tracker generated successfully!")
            st.download_button(
                label="📥 Download Excel Tracker",
                data=excel_data,
                file_name=f"attendance_tracker_{start_input.year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error during processing: {e}")